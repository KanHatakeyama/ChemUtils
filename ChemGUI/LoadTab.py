import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from .SaveSettings import SaveSettings
import pandas as pd
from tkinter import messagebox
from openpyxl import load_workbook

settings = SaveSettings()


class LoadTab(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.file_path_excel = tk.StringVar(value=settings.load("file_path_excel"))

        self.select_file_btn_excel = tk.Button(
            self, text="Select Excel path", command=self.select_file_excel
        )
        self.select_file_btn_excel.pack(side="top")

        self.file_path_entry_excel = tk.Entry(
            self, textvariable=self.file_path_excel, width=200
        )
        self.file_path_entry_excel.pack(side="top")

        self.load_button = tk.Button(
            self, text="Preload", command=self.on_button_click_load
        )
        self.load_button.pack()

    # ダイアログ関連
    def select_file_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file_path_excel.set(file_path)
        settings.set("file_path_excel", file_path)

    def on_button_click_load(self):
        try:
            self.load_excel()
            # messagebox.showinfo("success", "loaded excel")

        except Exception as e:
            messagebox.showinfo("error", "Load error" + str(e))

    def load_excel(self):
        self.workbook = load_workbook(filename=self.file_path_excel.get())
        sheet_names = self.workbook.sheetnames

        # ラベルを作成する
        label = tk.Label(self, text="Select a sheet:", font=("Arial", 12))
        label.pack(pady=10)

        # タブを選択するコンボボックスを作成する
        self.combo = ttk.Combobox(self, values=sheet_names, state="readonly")
        self.combo.set(sheet_names[0])
        self.combo.pack()

        # ボタンを作成する
        self.sheet_select_button = tk.Button(
            self, text="OK", command=self.on_button_set_sheet
        )
        self.sheet_select_button.pack()

    def on_button_set_sheet(self):
        sheetname = self.combo.get()
        file_path_excel = self.file_path_excel.get()
        label = tk.Label(
            self,
            text=f"{file_path_excel} \n {sheetname} will be processed",
            font=("Arial", 12),
        )
        label.pack()
        ws = self.workbook[sheetname]
        data = ws.values
        columns = next(data)[0:]
        self.master.df = pd.DataFrame(data, columns=columns)

        # カラム名がNoneの時に修正する
        for i, name in enumerate(self.master.df.columns):
            if name == None:
                self.master.df.columns.values[i] = "Unnamed: " + str(i)

        # 他のタブをアンロックする
        self.master.unlock_tabs()
